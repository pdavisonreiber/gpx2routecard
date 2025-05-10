#!/usr/bin/env python3

import sys
import os
import pandas as pd
import gpxpy
from OSGridConverter import latlong2grid
import math
import openpyxl
import requests

def haversine(lat1, lon1, lat2, lon2):
	R = 6371.0
	phi1 = math.radians(lat1)
	phi2 = math.radians(lat2)
	dphi = math.radians(lat2 - lat1)
	dlambda = math.radians(lon2 - lon1)
	a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
	c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
	return R * c

def calculate_bearing(lat1, lon1, lat2, lon2):
	# Convert to radians
	lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
	
	# Calculate bearing
	y = math.sin(lon2 - lon1) * math.cos(lat2)
	x = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(lon2 - lon1)
	bearing = math.degrees(math.atan2(y, x))
	
	# Convert to 0-360 range
	bearing = (bearing + 360) % 360
	
	# Round to nearest degree
	return round(bearing)

def extract_named_waypoints(gpx_file):
	with open(gpx_file, 'r') as f:
		gpx = gpxpy.parse(f)
	waypoints = []
	route_points = []
	
	# Get points from either route or track
	if gpx.routes and gpx.routes[0].points:
		for pt in gpx.routes[0].points:
			route_points.append((pt.latitude, pt.longitude))
	elif gpx.tracks and gpx.tracks[0].segments and gpx.tracks[0].segments[0].points:
		for pt in gpx.tracks[0].segments[0].points:
			route_points.append((pt.latitude, pt.longitude))
	
	named_points = []
	if route_points:
		start = route_points[0]
		grid_ref = latlong2grid(start[0], start[1])
		named_points.append(('START', '', start[0], start[1], grid_ref))
	for wpt in gpx.waypoints:
		if wpt.name:
			grid_ref = latlong2grid(wpt.latitude, wpt.longitude)
			named_points.append((wpt.name, wpt.description or '', wpt.latitude, wpt.longitude, grid_ref))
	# Rename the last waypoint to END
	if named_points:
		last_point = named_points[-1]
		named_points[-1] = ('END', last_point[1], last_point[2], last_point[3], last_point[4])
	
	# Get elevations for all waypoints
	locations = [{"latitude": lat, "longitude": lon} for _, _, lat, lon, _ in named_points]
	elevations = get_elevations_post(locations)
	
	def nearest_idx(lat, lon):
		return min(range(len(route_points)), key=lambda i: haversine(lat, lon, route_points[i][0], route_points[i][1]))
	indices = [nearest_idx(lat, lon) for (_, _, lat, lon, _) in named_points]
	distances = []
	bearings = []
	ascents = []  # List to store ascent/descent values
	
	# Add empty bearing and ascent for first waypoint
	bearings.append('')
	ascents.append('')
	
	# Calculate distances between consecutive waypoints
	for i in range(len(indices)-1):
		total = 0.0
		for j in range(indices[i], indices[i+1]):
			lat1, lon1 = route_points[j]
			lat2, lon2 = route_points[j+1]
			total += haversine(lat1, lon1, lat2, lon2)
		distances.append(round(total, 2))
		
		# Calculate bearing from previous waypoint to current waypoint
		bearing = calculate_bearing(named_points[i][2], named_points[i][3], 
								  named_points[i+1][2], named_points[i+1][3])
		bearings.append(f"{bearing:03d}°")
		
		# Calculate elevation change (positive for ascent, negative for descent)
		elev_diff = elevations[i+1][2] - elevations[i][2]
		ascents.append(round(elev_diff))  # Show both positive and negative values
	
	# Add empty values for last waypoint
	bearings.append('')
	ascents.append('')
	
	for i, (name, desc, lat, lon, grid_ref) in enumerate(named_points):
		waypoints.append({
			'grid_reference': str(grid_ref).rjust(8),  # Right-align grid references
			'name': name,
			'distance_from_last_km': distances[i-1] if i > 0 else '',  # Empty for START, distance for others
			'bearing': bearings[i],
			'description': desc,
			'ascent': ascents[i]  # Add ascent/descent in meters
		})
	return waypoints

def get_elevations_post(locations):
	url = "https://api.open-elevation.com/api/v1/lookup"
	response = requests.post(url, json={"locations": locations})
	response.raise_for_status()
	results = response.json()["results"]
	return [(r["latitude"], r["longitude"], r["elevation"]) for r in results]

def main():
	if len(sys.argv) != 2:
		print(f"Usage: {os.path.basename(sys.argv[0])} INPUT.gpx")
		sys.exit(1)
	inp = sys.argv[1]
	base, ext = os.path.splitext(inp)
	if ext.lower() != ".gpx":
		print(f"Error: expected a .gpx file, got {ext!r}")
		sys.exit(1)
	outp = base + ".xlsx"
	try:
		waypoints = extract_named_waypoints(inp)
		if not waypoints:
			print("No waypoints found in the GPX file", file=sys.stderr)
			sys.exit(1)
		df = pd.DataFrame(waypoints)
		df = df[['grid_reference', 'name', 'distance_from_last_km', 'bearing', 'ascent', 'description']]
		
		# Add empty Escape Notes column
		df['escape_notes'] = ''
		
		# Set custom column headings
		df.columns = ['Grid Reference', 'Waypoint', 'Distance (km)', 'Bearing', 'Ascent (m)', 'Description', 'Escape Notes']
		
		with pd.ExcelWriter(outp, engine='openpyxl') as writer:
			df.to_excel(writer, index=False, header=True, sheet_name='Route Card')
			worksheet = writer.sheets['Route Card']
			
			# Set column widths
			worksheet.column_dimensions['A'].width = 15  # Grid Reference
			worksheet.column_dimensions['B'].width = 20  # Waypoint
			worksheet.column_dimensions['C'].width = 20  # Distance from Last
			worksheet.column_dimensions['D'].width = 10  # Bearing
			worksheet.column_dimensions['E'].width = 20  # Elevation Change
			worksheet.column_dimensions['F'].width = 15  # Time
			worksheet.column_dimensions['G'].width = 15  # Rest Time
			worksheet.column_dimensions['H'].width = 15  # Arrival Time
			worksheet.column_dimensions['I'].width = 60  # Description
			worksheet.column_dimensions['J'].width = 30  # Escape Notes
			
			# Insert Time column after Ascent
			worksheet.insert_cols(6)
			worksheet.cell(row=1, column=6, value='Time (min)')
			
			# Add time formula to each row
			for row in range(2, len(df) + 2):
				# Simple formula using cell references
				formula = f'=IF(AND(C{row}<>"",E{row}<>""),C{row}*20+E{row}/10,"")'
				worksheet.cell(row=row, column=6, value=formula)
			
			# Insert Rest Time column after Time
			worksheet.insert_cols(7)
			worksheet.cell(row=1, column=7, value='Rest (min)')
			
			# Add rest time formula to each row
			for row in range(2, len(df) + 2):
				if row == 2:
					# First row (START) gets empty
					formula = '=""'
				else:
					# All other rows get 10 minutes
					formula = '=10'
				worksheet.cell(row=row, column=7, value=formula)
			
			# Insert Arrival Time column after Rest Time
			worksheet.insert_cols(8)
			worksheet.cell(row=1, column=8, value='Arrival Time')
			
			# Add arrival time formula to each row
			for row in range(2, len(df) + 2):
				if row == 2:
					# First row (START) gets 08:00
					formula = '=TIME(8,0,0)'
				else:
					# Subsequent rows add the time and rest from current row to previous arrival time
					formula = f'=IF(F{row}<>"",H{row-1}+TIME(0,F{row}+G{row},0),"")'
				cell = worksheet.cell(row=row, column=8, value=formula)
				cell.number_format = 'hh:mm'  # Format as time
			
			# Create a table with formatting
			table = openpyxl.worksheet.table.Table(
				displayName="RouteCardTable",
				ref=f"A1:J{len(df) + 1}"
			)
			worksheet.add_table(table)
			
			# Add borders and alternating row colors
			thin_border = openpyxl.styles.Border(
				left=openpyxl.styles.Side(style='thin'),
				right=openpyxl.styles.Side(style='thin'),
				top=openpyxl.styles.Side(style='thin'),
				bottom=openpyxl.styles.Side(style='thin')
			)
			
			# Apply borders and alternating colors to all cells
			for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=10), 1):
				for cell in row:
					cell.border = thin_border
					if row_idx > 1 and row_idx % 2 == 0:  # Even rows after header
						cell.fill = openpyxl.styles.PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
			
			# Make header row bold and center-aligned with gray background
			header_fill = openpyxl.styles.PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
			for cell in worksheet[1]:
				cell.font = cell.font.copy(bold=True)
				cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				cell.fill = header_fill
			
			# Set cell alignments for data rows
			for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
				# Grid Reference - centered
				row[0].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				# Waypoint - wrapped and centered
				row[1].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='center')
				# Distance from Last - centered
				row[2].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				# Bearing - centered
				row[3].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				# Elevation Change - centered
				row[4].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				# Time - centered
				row[5].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				# Rest Time - centered
				row[6].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				# Arrival Time - centered
				row[7].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
				# Description - wrapped
				row[8].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
				# Escape Notes - wrapped
				row[9].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
		print(f"Converted {inp!r} → {outp!r}")
		print(f"Found {len(waypoints)} waypoints")
	except gpxpy.gpx.GPXXMLSyntaxException as e:
		print(f"Invalid GPX file: {e}", file=sys.stderr)
		sys.exit(1)
	except Exception as e:
		print(f"Conversion failed: {e}", file=sys.stderr)
		sys.exit(1)

if __name__ == "__main__":
	main() 